#!/usr/bin/env python3
import rospy
import serial
import  sys
import openpyxl
import threading
from std_msgs.msg import String
from delivery_robot_serial_handler.msg import Vel,Pos

rospy.init_node("SerialHandler_node")

pub_V = rospy.Publisher('Vel', Vel, queue_size=10)
pub_P = rospy.Publisher('Pos', Pos, queue_size=10)

msg_V = Vel()
msg_P = Pos()

rospy.loginfo("Hello from serial handler node")
wb = openpyxl.Workbook()
ws = wb.active
ws.cell( row = 1,column =1,value = "s_rate1")
ws.cell( row = 1,column =2,value = "rate1")
ws.cell( row = 1,column =3,value = "s_rate2")
ws.cell( row = 1,column =4,value = "rate2")
i= 1
s_rate1 =0
s_rate2 =0
s_pos1 =0
s_pos2 =0
mess = ""
# initialize UART
ser = serial.Serial( port="/dev/ttyS0", baudrate=115200) #/dev/ttyAMA0   /dev/ttyUSB0 /dev/ttyS0


def callback(command):
    global s_rate1,s_rate2
    rospy.loginfo(command.data)
    command = str(command.data)
    data = command
    data = data.replace("!", "")
    data = data.replace("#", "")
    ser.write((command).encode())
    if (data[0] == 'V'):
        splitData = data.split(";")
        s_rate1 = int(splitData[0].replace("V",""))
        s_rate2 = int(splitData[1])
    elif (data[0] == 'P'):
        splitData = data.split(";")
        s_pos1 = str(splitData[0])
        s_pos2 = str(splitData[1])
def processData(data):
    data = data.replace("!", "")
    data = data.replace("#", "")
    splitData = data.split(";")
    #print(splitData)
    try:
        
        if splitData[0] == 'V':
            global i,s_rate1,s_rate2
            i = i +1
            msg_V.V1 = int(splitData[1])
            msg_V.V2 = int(splitData[2])
            pub_V.publish(msg_V)
            ws.cell(row = i, column = 2,value= msg_V.V1)
            ws.cell(row = i, column = 4,value= msg_V.V2)
            ws.cell(row = i, column = 1, value =  s_rate1)
            ws.cell(row = i, column = 3, value =  s_rate2)
        elif splitData[0] =='P':
            msg_P.P1 = int(splitData[1])
            msg_P.P2 = int(splitData[2])
            pub_P.publish(msg_P)
    except:
        pass

def readSerial():
    bytesToRead = ser.inWaiting()
    if (bytesToRead > 0):
        global mess
        try:
            mess = mess + ser.read(bytesToRead).decode("UTF-8")
            while ("#" in mess) and ("!" in mess):
                start = mess.find("!")
                end = mess.find("#")
                processData(mess[start:end + 1])
                if (end == len(mess)):
                    mess = ""
                else:
                    mess = mess[end+1:]
        except :
            print("Noise ! \n")
    #print(mess)
  
            
# main
rospy.Subscriber('command', String, callback)

        
while not rospy.is_shutdown():
    readSerial()
    
wb.save("test2.xlsx")
rospy.loginfo ("end of program")