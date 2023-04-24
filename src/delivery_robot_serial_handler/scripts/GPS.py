#!/usr/bin/env python3
import rospy
from std_msgs.msg import String
import serial
import openpyxl
import time
import math
from Kalman import Kalman_filter
from delivery_robot_serial_handler.msg import Pos

# wb = openpyxl.Workbook()
# ws = wb.active
# ws.cell( row = 1,column =1,value = "Lat")
# ws.cell( row = 1,column =2,value = "lon")
# i = 2
rospy.init_node("GPS_node")
rospy.loginfo ("GPS node has been Started !")

pub = rospy.Publisher('decition', String, queue_size=10)

P1 =0
P2 =0
P1_pre = 0
P2_pre = 0
def enc_callback(msg):

    global P1, P2   
    P1 = msg.P1
    P2 = msg.P2


def haversine(lat1, lon1, lat2, lon2):
     
    # distance between latitudes
    # and longitudes
    dLat = (lat2 - lat1) * math.pi / 180.0
    dLon = (lon2 - lon1) * math.pi / 180.0
 
    # convert to radians
    global lat1_r
    lat1_r = (lat1) * math.pi / 180.0
    lat2_r = (lat2) * math.pi / 180.0
 
    # apply formulae
    a = (pow(math.sin(dLat / 2), 2) +
         pow(math.sin(dLon / 2), 2) *
             math.cos(lat1_r) * math.cos(lat2_r));
    rad = 6371
    c = 2 * math.asin(math.sqrt(a))
    return rad * c * 100000
x0 = 60
khoang_cach = 0
estimate_dis = 0
Kal = Kalman_filter(11000, 25, 1000000000)
toado = ""
def parseGPS(data):
#     global i
    global Kal,x0,P1_pre, P2_pre, P1, P2
    if data[0:6] == b'$GPGGA':
        data = str(data)
        s = data.split(",")
        if s[6] == '0':
            rospy.loginfo("no satellite data")
            return 0
        lat = decode(s[2])
        lon = decode(s[4])
        lat_str = str(lat)
        lon_str = str(lon)
        global estimate_dis
        khoang_cach = haversine(lat,lon,10.88073469,106.8092379)
        estimate_dis = Kal.Update(x0,khoang_cach)
        x0 = khoang_cach - (P1 -P1_pre + P2- P2_pre)*40/4830
        P1_pre = P1
        P2_pre = P2
#         ws.cell( row = i ,column =1,value = lat)
#         ws.cell( row = i,column =2,value = lon)
        global toado
        toado = "&lat="+lat_str+"&lon="+lon_str
        #ver.write(toado.encode())
#         i = i+1
#         ver.write(toado.encode())
#         ver.flush()
#         print(toado)
#        time.sleep(4)
        return 1
    else:
        return 0        
def decode(coord):
    v = coord.split(".")
    head = v[0]
    tail = "0."+v[-1]
    deg = head[0:-2]
    minn = head[-2:]
#     return deg + minn + "." + tail
    decimal = int(deg) + (int(minn) + float(tail))/60
    return decimal
    
ser = serial.Serial(
        port = "/dev/ttyUSB0",
        baudrate = 9600,
        timeout = 1
    )

# ver = serial.Serial(
#        port = "/dev/ttyUSB1",
#        baudrate = 115200,
#        timeout = 1
#    )
# count =2

rospy.Subscriber('Pos', Pos, enc_callback)
rate = rospy.Rate(0.5)
while not rospy.is_shutdown():
    data = ser.readline()
    if parseGPS(data) :
        rospy.loginfo(toado)
        rospy.loginfo("estimate dis "+ str(estimate_dis))
        #rospy.loginfo(estimate_dis)
        if estimate_dis < 200:
            command = 'R'
        else:
            command = 'S'
        rospy.loginfo("enc "+str(P1 - P1_pre))
        P1_pre = P1
        pub.publish(command)
    #rate.sleep()
#     ver.write(toado)
#     ver.flush()
#     print(toado)
#    time.sleep(1)
# wb.save("GPS.xlsx")   
    
    
